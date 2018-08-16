#新浪并没有开放关键字搜索的API，所以采用微博的高级搜索功能进行爬虫
import urllib
from urllib import request
from urllib import parse
import time
import logging
from lxml import etree
import random
from datetime import datetime
from datetime import timedelta
import openpyxl


class CollectData():
    def __init__(self,keyword,startTime,path,session,interval='50',flag=True,begin_url_per="http://s.weibo.com/weibo/"):
        self.begin_url_per=begin_url_per #设置固定地址部分
        self.startTime=startTime
        self.session=session
        self.path = path
        self.keyWord=keyword
        self.setKeyWord(keyword)
        self.setStartTimescope(startTime)
        self.setInterval(interval)
        self.setFlag(flag)
        self.logger=logging.getLogger('main.CollectData')
        self.setExcel()
        self.num=0#当前爬虫条数


    def setExcel(self):
        self.wb=openpyxl.load_workbook(self.path)
        #title=self.startTime #sheet名字
        title=self.startTime+"-"+self.keyWord
        sheets=self.wb.sheetnames
        for i in range(len(sheets)):
            if sheets[i]==title:
                self.wb.remove(self.wb[title]) #如果存在旧表，则删除
                break

        self.sheet=self.wb.create_sheet(title)
        self.sheet.append(['序号','昵称','关键词','发表时间','微博地址','微博内容'])

    ##设置关键字
    ##关键字需解码后编码为utf-8
    def setKeyWord(self,keyword):
        self.keyword=keyword.encode("utf-8")
        print('twice encode:',self.getKeyWord())

    ##关键词需要进行两次urlencode
    def getKeyWord(self):
        once=urllib.parse.urlencode({"kw":self.keyword})[3:]
        return urllib.parse.urlencode({"kw":once})[3:]

    ##设置起始范围，间隔为1天
    ##格式为：yyyy-mm-dd
    def setStartTimescope(self,startTime):
        if not (startTime=='-'):
            self.timescope=startTime+":"+startTime
        else:
            self.timescope='-'

    ##设置搜索地区
    def setRegion(self,region):
        self.region=region

    ##设置邻近网页请求之间的基础间隔时间
    def setInterval(self,interval):
        self.interval=int(interval)

    ##设置是否被认为机器人的标志，若为False，需要进入页面，手动输入验证码
    def setFlag(self,flag):
        self.flag=flag

    ##构建URL
    def getURL(self):
        return self.begin_url_per+self.getKeyWord()+"&typeall=1&suball=1&timescope=custom:"+self.timescope+"&page="

    ##爬取一次请求中的所有网页，最多返回50页
    def download(self,url,maxTryNum=3):
        hasMore=True #某次请求可能少于50页，设置标记，判断是否还有下一页
        isCaught=False #某次请求被认为是机器人，设置标记，判断是否被抓住。抓住后，需要，进入页面，输入验证码
        name_filter=set([]) #过滤重复的微博ID

        i=1 #记录本次请求所返回的页数
        while hasMore and i<9 and (not isCaught):#最多返回50页，对每页进行解析
            source_url=url+str(i) #构建某页的URL
            data='' #存储该页的网页数据
            goon=True #网络中断标记
            print('当前页： '+str(i))
            ##网络不好的情况，试着尝试请求三次
            for tryNum in range(maxTryNum):
                try:

                    #html=urllib.request.urlopen(source_url,timeout=12)
                    #session=requests.session()
                    html=self.session.get(source_url)#必须带着登录cookie
                    data=html.text.encode('utf-8')
                    #data=html.read()
                    break
                except:
                    if tryNum<(maxTryNum-1):
                        time.sleep(10)
                    else:
                        print('Internet Connect Error!')
                        self.logger.error('Internet Connect Error!')
                        self.logger.info('url: '+source_url)
                        self.logger.info('page: '+str(i))
                        self.flag=False
                        goon=False
                        break
            if goon:
                lines=data.splitlines()
                isCaught=True
                for line in lines:
                    ## 判断是否有微博内容，出现这一行，则说明没有被认为是机器人
                    if line.startswith(b'<script>STK && STK.pageletM && STK.pageletM.view({"pid":"pl_weibo_direct"'):
                        isCaught=False
                        n=line.find(b'html":"')
                        if n>0:
                            j=line[n+7:-12].decode('unicode_escape').replace("\\","")
                            ##如果没有更多结果页面
                            #if (j.find("//div[@class='pl_noresult']") > 0):
                            if(j.find('<div class="search_noresult">')>0):
                                print('没有结果页了')
                                hasMore=False
                            ##有结果的页面
                            else:
                                ##此处j要decode，因为上面j被encode成utf-8了
                                page=etree.HTML(j)
                                ps=page.xpath("//p[@node-type='feed_list_content']")   #使用xpath解析得到微博内容
                                addrs=page.xpath("//a[@class='W_texta W_fb']")
                                addri=0
                                #获取昵称和微博内容
                                for p in ps:
                                    name=p.attrib.get('nick-name') #获取昵称
                                    txt=p.xpath('string(.)') #获取微博内容
                                    addr=addrs[addri].attrib.get('href') #获取微博地址
                                    addri+=1
                                    if(name!='None' and str(txt)!='None' and name not in name_filter):#导出数据到excel
                                        name_filter.add(name)
                                        self.num=self.sheet.max_row#获取当前数目
                                        self.sheet.append([self.num,name,self.keyWord,self.startTime,addr,txt])
                                        ##保存表
                                        self.wb.save(self.path)
                        break

                lines=None
                ##处理被认为是机器人的情况
                if isCaught:
                    print('Be Caught!')
                    self.logger.error('Be Caught Error!')
                    self.logger.info('url: '+source_url)
                    self.logger.info('page: '+str(i))
                    data=None
                    self.flag=False
                    break
                ##没有更多结果，结束该次请求，跳到下一个请求
                if not hasMore:
                    print('No More Results')
                    if i==1:
                        time.sleep(random.randint(3,8))
                    else:
                        time.sleep(10)
                    data=None
                    break
                i+=1
                ##设置两个临近URL请求之间的随机休眠时间，防止Be Caught
                sleeptime_one=random.randint(self.interval-35,self.interval-15)
                sleeptime_two=random.randint(self.interval-15,self.interval+10)
                #sleeptime_one = random.randint(0, 3)
                #sleeptime_two = random.randint(4, 6)
                if i%2==0:
                    sleeptime=sleeptime_two
                else:
                    sleeptime=sleeptime_one

                print('sleeping '+str(sleeptime)+' seconds...')
                time.sleep(sleeptime)
            else:
                break

    ##改变搜索的时间范围，有利于获取更多的数据，搜索到前一天
    def getTimescope(self,perTimescope):
        if not (perTimescope=='-'):
            times_list = perTimescope.split(':')
            start_date = datetime(int(times_list[-1][0:4]), int(times_list[-1][5:7]), int(times_list[-1][8:10]))
            start_new_date = start_date - timedelta(days=1)
            start_str = start_new_date.strftime("%Y-%m-%d")
            return start_str + ":" + start_str
        else:
            return "-"

    def start(self):
        while self.num<800 and self.flag:
            print("当前时间为：%s"%self.timescope)
            url=self.getURL()
            self.download(url=url)
            self.timescope=self.getTimescope(self.timescope) #改变搜索时间，到前一天

        print("爬虫结束，共得到 %d 条数据"%self.num)

    def main(self):
        logger = logging.getLogger('main')
        logFile = './collect.log'
        logger.setLevel(logging.DEBUG)
        filehandler = logging.FileHandler(logFile)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s: %(message)s')
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)

if __name__=="__main__":

    '''
    logger = logging.getLogger('main')
    logFile = './collect.log'
    logger.setLevel(logging.DEBUG)
    filehandler = logging.FileHandler(logFile)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s: %(message)s')
    filehandler.setFormatter(formatter)
    logger.addHandler(filehandler)
    '''
    keyword='小米'
    startTime='2018-08-13'
    interval='40'
    path='data/weibo.xlsx'
    cd=CollectData(keyword,startTime,path,interval)
    print(cd.getTimescope(startTime))

    #url=cd.getURL()
    #cd.download(url)






